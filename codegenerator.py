import streamlit as st
import qrcode
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import pandas as pd
import zipfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

st.title("QR Code & Barcode Generator")

tab1, tab2 = st.tabs(["Single", "Batch"])

def generate_code(text, code_type):
    buffer = BytesIO()
    if code_type == "QR Code":
        img = qrcode.make(str(text))
        img.save(buffer, format="PNG")
    else:
        type_map = {
            "Code128": "code128",
            "Code39": "code39",
            "EAN13": "ean13",
            "EAN8": "ean8",
            "UPCA": "upca",
        }
        bc_class = barcode.get_barcode_class(type_map[code_type])
        bc = bc_class(str(text), writer=ImageWriter())
        bc.write(buffer)
    buffer.seek(0)
    return buffer

with tab1:
    text = st.text_input("Enter Text", "ER46")
    code_type = st.selectbox("Code Type", ["QR Code", "Code128", "Code39", "EAN13", "EAN8", "UPCA"])

    if st.button("Generate", key="single_generate"):
        try:
            buf = generate_code(text, code_type)
            st.image(buf, caption=f"{code_type}: {text}")
            st.download_button(
                "Download",
                data=buf.getvalue(),
                file_name=f"{text}_{code_type}.png",
                mime="image/png"
            )
        except Exception as e:
            st.error(f"Error: {e}")

with tab2:
    with st.expander("📋 File Format Guide", expanded=True):
        st.markdown("""
**Required column:** `text` — the value to encode. The output filename will use this value.  
Any extra columns are ignored.

**Character limits by code type:**
| Code Type | `text` requirement |
|-----------|-------------------|
| QR Code | 任意字符串（含中文） |
| Code128 | 任意 ASCII 字符串（字母、数字、符号，不含中文） |
| Code39 | 大写字母、数字，以及 - . $ / + % 空格 |
| EAN13 | **12** 位数字（末位校验码自动生成） |
| EAN8 | **7** 位数字（末位校验码自动生成 |
| UPCA | **11** 位数字（末位校验码自动生成） |
""")
        template_df = pd.DataFrame({"text": ["ER46", "ER47", "ER48"]})
        st.dataframe(template_df, use_container_width=True, hide_index=True)
        st.download_button(
            "⬇️ Download template CSV",
            data=template_df.to_csv(index=False).encode(),
            file_name="batch_upload_template.csv",
            mime="text/csv",
        )

    uploaded = st.file_uploader("Upload an Excel or CSV file", type=["xlsx", "xls", "csv"])

    if uploaded:
        try:
            df = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
        except Exception as e:
            st.error(f"Failed to read file: {e}")
            st.stop()

        st.dataframe(df.head(), use_container_width=True)

        col1, col2 = st.columns(2)
        with col1:
            text_col = st.selectbox("Text Column", df.columns.tolist())
        with col2:
            batch_type = st.selectbox("Code Type", ["QR Code", "Code128", "Code39", "EAN13", "EAN8", "UPCA"], key="batch")

        output_format = st.radio("Output Format", ["ZIP (individual PNGs)", "Excel (images in column B)"], horizontal=True)

        if st.button("Generate", key="batch_generate"):
            values = df[text_col].dropna().tolist()
            if not values:
                st.warning("No valid data in the selected column.")
                st.stop()

            errors = []
            progress = st.progress(0, text="Generating...")

            # ── Excel output ──────────────────────────────
            if output_format == "Excel (images in column B)":
                wb = Workbook()
                ws = wb.active
                ws.column_dimensions["A"].width = 20
                ws.column_dimensions["B"].width = 40

                # Header
                ws["A1"] = "text"
                ws["B1"] = batch_type

                ROW_HEIGHT = 80  # pixels → approx row height in points
                IMG_H = 70       # image height to embed (px)

                for i, val in enumerate(values):
                    row = i + 2
                    ws.row_dimensions[row].height = ROW_HEIGHT
                    ws[f"A{row}"] = str(val)
                    try:
                        img_buf = generate_code(val, batch_type)
                        xl_img = XLImage(img_buf)
                        # Scale to fixed height, keep aspect ratio
                        ratio = IMG_H / xl_img.height
                        xl_img.height = IMG_H
                        xl_img.width = int(xl_img.width * ratio)
                        ws.add_image(xl_img, f"B{row}")
                    except Exception as e:
                        errors.append(f"Row {i+1} ({val}): {e}")
                    progress.progress((i + 1) / len(values), text=f"{i+1} / {len(values)}")

                xl_buf = BytesIO()
                wb.save(xl_buf)
                xl_buf.seek(0)

                st.success(f"Done! {len(values) - len(errors)} generated, {len(errors)} failed.")
                if errors:
                    with st.expander("View errors"):
                        for err in errors:
                            st.text(err)

                st.download_button(
                    "📊 Download Excel",
                    data=xl_buf.getvalue(),
                    file_name=f"batch_{batch_type}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # ── ZIP output ────────────────────────────────
            else:
                zip_buf = BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for i, val in enumerate(values):
                        try:
                            img_buf = generate_code(val, batch_type)
                            zf.writestr(f"{val}_{batch_type}.png", img_buf.getvalue())
                        except Exception as e:
                            errors.append(f"Row {i+1} ({val}): {e}")
                        progress.progress((i + 1) / len(values), text=f"{i+1} / {len(values)}")

                zip_buf.seek(0)
                st.success(f"Done! {len(values) - len(errors)} generated, {len(errors)} failed.")
                if errors:
                    with st.expander("View errors"):
                        for err in errors:
                            st.text(err)

                st.download_button(
                    "📦 Download ZIP",
                    data=zip_buf.getvalue(),
                    file_name=f"batch_{batch_type}.zip",
                    mime="application/zip"
                )